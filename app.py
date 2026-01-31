"""
Reshape Wizard MVP with preview-first UX.
- Load CSV/XLSX, optional unmerge+fill, fill-down, longer/wider/transpose.
- Every operation offers preview (before/after) before committing.
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import streamlit.components.v1 as components


PREVIEW_DEFAULT = 100
LARGE_ROW_THRESHOLD = 200_000
VERY_WIDE_THRESHOLD = 500
HUGE_ROW_WARNING = 5_000_000


def init_state() -> None:
    """Ensure session_state keys exist."""
    ss = st.session_state
    ss.setdefault("original_df", None)
    ss.setdefault("current_df", None)
    ss.setdefault("history", [])
    ss.setdefault("max_history", 20)
    ss.setdefault("last_merge_fill", None)
    ss.setdefault("preview_n", PREVIEW_DEFAULT)
    ss.setdefault("preview_df", None)
    ss.setdefault("preview_meta", None)
    ss.setdefault("preview_error", None)
    ss.setdefault("preview_params_hash", None)
    ss.setdefault("preview_limit", PREVIEW_DEFAULT)
    ss.setdefault("upload_cache", {})  # {"bytes":..., "name":..., "ext":..., "sheets": [...]}


def normalise_duplicate_columns(cols) -> List[str]:
    """Suffix duplicate column names with .2, .3 ..."""
    counts: Dict[str, int] = {}
    new_cols: List[str] = []
    for col in cols:
        base = "Unnamed" if pd.isna(col) else str(col)
        counts[base] = counts.get(base, 0) + 1
        new_cols.append(base if counts[base] == 1 else f"{base}.{counts[base]}")
    return new_cols


def is_invalid_colname(name: Any) -> bool:
    if pd.isna(name):
        return True
    if isinstance(name, str):
        stripped = name.strip()
        if stripped == "":
            return True
        if re.match(r"^Unnamed(:\s*\d+)?$", stripped, re.IGNORECASE):
            return True
    return False


def sanitise_columns(columns: List[Any]) -> Tuple[List[str], bool, List[Tuple[Any, str]]]:
    """Return safe column names, whether fixes applied, and mapping."""
    provisional: List[str] = []
    fixed = False
    mapping: List[Tuple[Any, str]] = []
    for idx, col in enumerate(columns, start=1):
        if is_invalid_colname(col):
            suggested = f"Column_{idx}"
            fixed = True
        else:
            suggested = str(col).strip()
        provisional.append(suggested)
        mapping.append((col, suggested))
    deduped = normalise_duplicate_columns(provisional)
    if deduped != provisional:
        fixed = True
        mapping = [(orig, new) for (orig, _), new in zip(mapping, deduped)]
    return deduped, fixed, mapping


def add_history_step(name: str, params: Dict[str, Any], df: pd.DataFrame, before_shape: Tuple[int, int]) -> None:
    """Record a transformation; keep max_history."""
    step = {
        "name": name,
        "params": params,
        "before_shape": before_shape,
        "after_shape": df.shape,
        "df": df.copy(deep=True),
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }
    st.session_state.history.append(step)
    if len(st.session_state.history) > st.session_state.max_history:
        st.session_state.history.pop(0)


def clear_preview() -> None:
    st.session_state.preview_df = None
    st.session_state.preview_meta = None
    st.session_state.preview_error = None
    st.session_state.preview_params_hash = None


def get_sheet_names(file_bytes: bytes) -> List[str]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    return wb.sheetnames


def read_csv_with_fallback(file_bytes: bytes) -> pd.DataFrame:
    for enc in ("utf-8", "cp1251"):
        try:
            df = pd.read_csv(BytesIO(file_bytes), encoding=enc)
            df.columns = normalise_duplicate_columns(df.columns)
            return df
        except Exception:
            continue
    raise ValueError("Could not read CSV with utf-8 or cp1251 encodings.")


def unmerge_fill_xlsx(file_bytes: bytes, sheet_name: str, header_row: int) -> Tuple[pd.DataFrame, int]:
    """Fill merged cell ranges with their top-left value and return DataFrame."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    ws = wb[sheet_name]

    filled_ranges = 0
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left = ws.cell(row=min_row, column=min_col).value
        if top_left is None:
            ws.unmerge_cells(str(merged_range))
            continue
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left
        ws.unmerge_cells(str(merged_range))
        filled_ranges += 1

    data = [list(row) for row in ws.iter_rows(values_only=True)]
    if not data:
        return pd.DataFrame(), filled_ranges

    header_idx = max(header_row - 1, 0)
    if header_idx >= len(data):
        raise ValueError("Header row is beyond the data range.")

    header = data[header_idx]
    records = data[header_idx + 1 :]
    df = pd.DataFrame(records, columns=header)
    df.columns = normalise_duplicate_columns(df.columns)
    return df, filled_ranges


def load_xlsx_standard(file_bytes: bytes, sheet_name: str, header_row: int) -> pd.DataFrame:
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=header_row - 1, engine="openpyxl")
    df.columns = normalise_duplicate_columns(df.columns)
    return df


def load_dataframe(
    file_bytes: bytes,
    filename: str,
    sheet_name: Optional[str],
    header_row: int,
    unmerge_fill: bool,
) -> Tuple[pd.DataFrame, Optional[int]]:
    ext = filename.lower().split(".")[-1]
    merge_info: Optional[int] = None
    if ext == "csv":
        df = read_csv_with_fallback(file_bytes)
    elif ext in ("xlsx", "xlsm", "xls"):
        if sheet_name is None:
            raise ValueError("Please choose a sheet.")
        if unmerge_fill:
            df, merge_info = unmerge_fill_xlsx(file_bytes, sheet_name, header_row)
        else:
            df = load_xlsx_standard(file_bytes, sheet_name, header_row)
    else:
        raise ValueError("Unsupported file type. Please upload CSV or XLSX.")

    if df.empty or df.shape[1] < 1:
        raise ValueError("Loaded data is empty or has no columns.")
    return df, merge_info


def load_from_clipboard(text: str) -> pd.DataFrame:
    """Parse tab-separated clipboard text from Excel/Sheets."""
    cleaned = text.strip()
    if not cleaned:
        raise ValueError("No data pasted.")
    try:
        df = pd.read_csv(StringIO(cleaned), sep="\t", dtype=str)
    except Exception as exc:
        raise ValueError(
            "Could not parse pasted table. Make sure it is copied from Excel as a table."
        ) from exc
    df = df.replace("", pd.NA)
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if df.empty or df.shape[1] == 0:
        raise ValueError("Pasted table is empty after trimming blank rows/columns.")

    safe_cols, _, _ = sanitise_columns(list(df.columns))
    df.columns = safe_cols
    return df


def fill_down(df: pd.DataFrame, columns: List[str], treat_empty_as_missing: bool) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """Fill down selected columns; return df and counts per column."""
    if not columns:
        return df, {}
    filled = df.copy()
    filled_counts: Dict[str, int] = {}
    for col in columns:
        before_na = filled[col].isna().sum()
        if treat_empty_as_missing:
            filled[col] = filled[col].replace("", pd.NA)
        filled[col] = filled[col].ffill()
        after_na = filled[col].isna().sum()
        filled_counts[col] = int(before_na - after_na)
    return filled, filled_counts


def make_longer(
    df: pd.DataFrame,
    id_cols: List[str],
    value_cols: List[str],
    name_col: str,
    value_col: str,
) -> pd.DataFrame:
    return pd.melt(df, id_vars=id_cols, value_vars=value_cols, var_name=name_col, value_name=value_col)


def make_wider(
    df: pd.DataFrame,
    index_cols: List[str],
    column_source: str,
    value_source: str,
    aggfunc: Optional[str],
    reset_index: bool,
) -> pd.DataFrame:
    work_df = df.copy()
    if aggfunc == "mean":
        work_df[value_source] = pd.to_numeric(work_df[value_source], errors="coerce")

    if aggfunc:
        widened = work_df.pivot_table(
            index=index_cols,
            columns=column_source,
            values=value_source,
            aggfunc=aggfunc,
        )
    else:
        widened = work_df.pivot(index=index_cols, columns=column_source, values=value_source)

    if isinstance(widened.columns, pd.MultiIndex):
        widened.columns = ["_".join(map(str, col)).strip() for col in widened.columns.values]
    else:
        widened.columns = [str(c) for c in widened.columns]

    if reset_index:
        widened = widened.reset_index()

    return widened


def transpose_df(
    df: pd.DataFrame,
    use_first_row_as_header: bool,
    use_first_col_as_index: bool,
) -> Tuple[pd.DataFrame, bool]:
    temp = df.copy()
    if use_first_col_as_index and not temp.empty:
        temp = temp.set_index(temp.columns[0])
    transposed = temp.transpose()

    header_adjusted = False
    if use_first_row_as_header and not transposed.empty:
        header = transposed.iloc[0]
        transposed = transposed.iloc[1:]
        transposed.columns = header
        header_adjusted = True

    transposed = transposed.reset_index().rename(columns={"index": "index"})
    before_cols = list(transposed.columns)
    transposed.columns = normalise_duplicate_columns(transposed.columns)
    warnings_adjusted = before_cols != list(transposed.columns)
    return transposed, header_adjusted or warnings_adjusted


def preview_rows(df: pd.DataFrame) -> int:
    limit = st.session_state.get("preview_limit", PREVIEW_DEFAULT)
    if df.shape[0] > LARGE_ROW_THRESHOLD:
        return min(50, df.shape[0], limit)
    return min(limit, df.shape[0])


def df_fingerprint(df: pd.DataFrame) -> str:
    """Lightweight fingerprint to key cache."""
    sample = df.head(50)
    values_hash = pd.util.hash_pandas_object(sample, index=True).sum()
    return f"{df.shape}-{hash(tuple(sample.columns))}-{int(values_hash)}"


def show_before_after(
    before_df: Optional[pd.DataFrame],
    after_df: Optional[pd.DataFrame],
    meta: Optional[Dict[str, Any]],
    error: Optional[str],
    title_after: str,
) -> None:
    col_before, col_after = st.columns(2)
    with col_before:
        if before_df is None:
            st.info("Load data to see a preview.")
        else:
            n = preview_rows(before_df)
            st.caption(f"Before • {before_df.shape[0]} rows × {before_df.shape[1]} columns (showing {n})")
            st.dataframe(before_df.head(n))
            st.caption("Dtypes")
            st.write(before_df.dtypes.to_frame("dtype"))
    with col_after:
        st.caption(title_after)
        if error:
            st.error(error)
            return
        if after_df is None or meta is None:
            st.info("Update preview to see the result.")
            return
        n = preview_rows(after_df)
        st.caption(f"After • {meta['after_shape'][0]} rows × {meta['after_shape'][1]} columns (showing {n})")
        st.dataframe(after_df.head(n))
        warnings = meta.get("warnings", [])
        diagnostics = meta.get("diagnostics", {})
        if warnings:
            st.warning("\n".join(warnings))
        if diagnostics:
            st.caption(f"Diagnostics: {diagnostics}")


def copy_to_clipboard_button(preview_df: Optional[pd.DataFrame], current_df: Optional[pd.DataFrame], label: str, key_prefix: str, include_index: bool = False, max_cells: int = 200_000) -> None:
    """Render a button that copies preview_df if present else current_df as TSV."""
    df = preview_df if preview_df is not None else current_df
    if df is None:
        st.button(label, key=f"copy_btn_{key_prefix}", disabled=True)
        return
    cells = df.shape[0] * df.shape[1]
    if cells > max_cells:
        st.warning(f"Copy limited to {max_cells} cells; table has {cells}. Please filter or export instead.")
        st.button(label, key=f"copy_btn_{key_prefix}", disabled=True)
        return
    tsv_text = df.to_csv(sep="\t", index=include_index)
    escaped = json.dumps(tsv_text)  # safe for JS string
    btn_html = f"""
        <button id="copy_{key_prefix}" style="padding:6px 10px;">{label}</button>
        <script>
        const btn = document.getElementById("copy_{key_prefix}");
        if (btn) {{
            btn.addEventListener("click", async () => {{
                try {{
                    await navigator.clipboard.writeText({escaped});
                    btn.innerText = "{label} ✓";
                    setTimeout(() => btn.innerText = "{label}", 1200);
                }} catch (err) {{
                    alert("Clipboard copy failed: " + err);
                }}
            }});
        }}
        </script>
    """
    components.html(btn_html, height=40)


def build_wider_work_df(df: pd.DataFrame, combined_parts: List[str], sep: str, column_source: Optional[str]) -> Tuple[pd.DataFrame, str]:
    """Return working df and the column to use for column names."""
    work_df = df.copy()
    column_used = column_source
    if combined_parts:
        for part in combined_parts:
            if part not in work_df.columns:
                raise ValueError(f"Column '{part}' not found for combination.")
        work_df["_combined_wider"] = work_df[combined_parts].astype(str).agg(sep.join, axis=1)
        column_used = "_combined_wider"
    if not column_used:
        raise ValueError("Choose column name parts or a single column.")
    if column_used not in work_df.columns:
        raise ValueError("Combined column was not created correctly.")
    return work_df, column_used


@st.cache_data(show_spinner=False, hash_funcs={pd.DataFrame: lambda _: 0})
def cached_preview(df_hash: str, op_name: str, params_json: str, df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    params = json.loads(params_json)
    return _compute_preview(df, op_name, params)


def _compute_preview(df: pd.DataFrame, op_name: str, params: Dict[str, Any]) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    before_shape = df.shape
    warnings: List[str] = []
    diagnostics: Dict[str, Any] = {}

    if op_name == "fill_down":
        columns = params.get("columns", [])
        treat_empty = params.get("treat_empty", True)
        if not columns:
            raise ValueError("Select at least one column to fill down.")
        df_out, counts = fill_down(df, columns, treat_empty)
        safe_cols, fixed, _ = sanitise_columns(list(df_out.columns))
        df_out.columns = safe_cols
        if fixed:
            warnings.append("Some column names were auto-corrected.")
        diagnostics["filled_counts"] = counts
        return df_out, {"before_shape": before_shape, "after_shape": df_out.shape, "warnings": warnings, "diagnostics": diagnostics}

    if op_name == "unmerge_fill":
        cache = st.session_state.upload_cache
        if not cache.get("bytes") or cache.get("name") != params.get("cache_key"):
            raise ValueError("No workbook available for unmerge preview. Reload the file first.")
        df_out, filled = unmerge_fill_xlsx(cache["bytes"], params["sheet_name"], params["header_row"])
        safe_cols, fixed, _ = sanitise_columns(list(df_out.columns))
        df_out.columns = safe_cols
        if fixed:
            warnings.append("Column names were auto-renamed after unmerge.")
        warnings.append(f"Filled {filled} merged ranges.")
        diagnostics["filled_ranges"] = filled
        return df_out, {"before_shape": before_shape, "after_shape": df_out.shape, "warnings": warnings, "diagnostics": diagnostics}

    if op_name == "longer":
        id_cols = params.get("id_cols", [])
        value_cols = params.get("value_cols", [])
        name_col = params.get("name_col") or "variable"
        value_col = params.get("value_col") or "value"
        if not value_cols:
            raise ValueError("Select at least one value column.")
        for col in id_cols + value_cols:
            if col not in df.columns:
                raise ValueError(f"Column '{col}' not found.")
        df_out = make_longer(df, id_cols, value_cols, name_col, value_col)
        safe_cols, fixed, _ = sanitise_columns(list(df_out.columns))
        df_out.columns = safe_cols
        if fixed:
            warnings.append("Some column names were auto-corrected.")
        if not id_cols:
            warnings.append("No ID columns selected; all columns stacked.")
        expected_rows = len(df) * len(value_cols)
        if expected_rows > HUGE_ROW_WARNING:
            warnings.append("Result is very large; applying may be slow.")
        return df_out, {"before_shape": before_shape, "after_shape": df_out.shape, "warnings": warnings, "diagnostics": diagnostics}

    if op_name == "wider":
        index_cols = params.get("index_cols", [])
        column_source = params.get("column_source")
        value_source = params.get("value_source")
        aggfunc = params.get("aggfunc")
        reset_index = params.get("reset_index", True)
        combined_parts = params.get("combined_parts", [])
        sep = params.get("sep", "_")

        work_df, column_used = build_wider_work_df(df, combined_parts, sep, column_source)

        if not value_source:
            raise ValueError("Choose a values column.")
        if column_used == value_source:
            raise ValueError("Column source and value source must differ.")
        for col in index_cols + [column_used, value_source]:
            if col not in work_df.columns:
                raise ValueError(f"Column '{col}' not found.")

        duplicates = work_df.duplicated(subset=index_cols + [column_used]).any()
        diagnostics["duplicates_detected"] = bool(duplicates)
        if duplicates and not aggfunc:
            raise ValueError("Duplicates detected; choose an aggregation.")
        coercions = 0
        if aggfunc == "mean":
            series = pd.to_numeric(work_df[value_source], errors="coerce")
            coercions = int(series.isna().sum() - work_df[value_source].isna().sum())
            work_df[value_source] = series
            df_out = make_wider(work_df, index_cols, column_used, value_source, aggfunc, reset_index)
        else:
            df_out = make_wider(work_df, index_cols, column_used, value_source, aggfunc, reset_index)
        if duplicates:
            warnings.append(f"Duplicates found; aggregation='{aggfunc}' applied.")
        if coercions > 0:
            warnings.append(f"Coerced {coercions} values to numeric for mean.")
            diagnostics["coercions"] = coercions
        if df_out.shape[1] > VERY_WIDE_THRESHOLD:
            warnings.append("Resulting table is very wide.")
        if combined_parts:
            diagnostics["combined_parts"] = combined_parts
            diagnostics["separator"] = sep
        safe_cols, fixed, _ = sanitise_columns(list(df_out.columns))
        df_out.columns = safe_cols
        if fixed:
            warnings.append("Some column names were auto-corrected.")
        return df_out, {"before_shape": before_shape, "after_shape": df_out.shape, "warnings": warnings, "diagnostics": diagnostics}

    if op_name == "transpose":
        use_first_row = params.get("first_row_header", True)
        use_first_col = params.get("first_col_index", False)
        df_out, adjusted = transpose_df(df, use_first_row, use_first_col)
        safe_cols, fixed, _ = sanitise_columns(list(df_out.columns))
        df_out.columns = safe_cols
        if adjusted:
            warnings.append("Headers were adjusted/normalised after transpose.")
        if fixed:
            warnings.append("Some column names were auto-corrected.")
        return df_out, {"before_shape": before_shape, "after_shape": df_out.shape, "warnings": warnings, "diagnostics": diagnostics}

    if op_name == "split_combined":
        cols_to_split = params.get("cols_to_split", [])
        sep = params.get("sep", "_")
        new_col1 = params.get("new_col1") or "part1"
        new_col2 = params.get("new_col2") or "part2"
        value_col = params.get("value_col") or "value"
        id_cols = params.get("id_cols", [])
        if not cols_to_split:
            raise ValueError("Select columns to split.")
        for c in cols_to_split:
            if c not in df.columns:
                raise ValueError(f"Column '{c}' not found.")
        # validate separator
        if any(sep not in c for c in cols_to_split):
            raise ValueError("Separator not found in all selected columns.")
        melt_id = id_cols if id_cols else [c for c in df.columns if c not in cols_to_split]
        df_melt = df.melt(id_vars=melt_id, value_vars=cols_to_split, var_name="combined", value_name=value_col)
        split = df_melt["combined"].str.split(sep, expand=True)
        if split.shape[1] != 2 or split.isna().any().any():
            raise ValueError("Splitting failed; inconsistent separator or parts.")
        split.columns = [new_col1, new_col2]
        df_out = pd.concat([df_melt.drop(columns=["combined"]), split], axis=1)
        safe_cols, fixed, _ = sanitise_columns(list(df_out.columns))
        df_out.columns = safe_cols
        if fixed:
            warnings.append("Some column names were auto-corrected.")
        diagnostics["split_separator"] = sep
        return df_out, {"before_shape": before_shape, "after_shape": df_out.shape, "warnings": warnings, "diagnostics": diagnostics}

    raise ValueError(f"Unknown operation '{op_name}'.")


def trigger_preview(df: pd.DataFrame, op_name: str, params: Dict[str, Any], auto_update: bool, params_hash: str, cache_df_hash: str) -> None:
    """Compute and store preview if needed."""
    do_update = auto_update and st.session_state.preview_params_hash != params_hash
    button_clicked = st.button("Update preview", key=f"update_{op_name}")
    if button_clicked:
        do_update = True
    if not do_update:
        return
    with st.spinner("Computing preview..."):
        try:
            preview_df, meta = cached_preview(cache_df_hash, op_name, json.dumps(params, sort_keys=True), df)
            st.session_state.preview_df = preview_df
            st.session_state.preview_meta = meta
            st.session_state.preview_error = None
            st.session_state.preview_params_hash = params_hash
        except Exception as exc:
            st.session_state.preview_df = None
            st.session_state.preview_meta = None
            st.session_state.preview_error = str(exc)
            st.session_state.preview_params_hash = params_hash


def get_preview_state(params_hash: str) -> Tuple[Optional[pd.DataFrame], Optional[Dict[str, Any]], Optional[str]]:
    """Return preview data only if it matches the provided hash."""
    if st.session_state.preview_params_hash == params_hash:
        return st.session_state.preview_df, st.session_state.preview_meta, st.session_state.preview_error
    return None, None, None


def render_history_sidebar() -> None:
    st.sidebar.subheader("History")
    history = list(reversed(st.session_state.history))
    if not history:
        st.sidebar.caption("No steps yet.")
        return

    for step in history:
        st.sidebar.write(
            f"{step['timestamp']} — {step['name']} "
            f"{step['before_shape']} → {step['after_shape']}"
        )

    col1, col2 = st.sidebar.columns(2)
    if col1.button("Undo last step", use_container_width=True):
        if len(st.session_state.history) > 1:
            st.session_state.history.pop()
            st.session_state.current_df = st.session_state.history[-1]["df"].copy()
            clear_preview()
        else:
            st.sidebar.warning("Nothing to undo.")

    if col2.button("Reset to original", use_container_width=True):
        if st.session_state.original_df is not None:
            st.session_state.current_df = st.session_state.original_df.copy()
            st.session_state.history = st.session_state.history[:1]
            clear_preview()
        else:
            st.sidebar.warning("No original data loaded.")


def export_buttons(df: pd.DataFrame) -> None:
    if df is None:
        st.sidebar.caption("Load data to enable export.")
        return
    st.sidebar.subheader("Export")
    include_index = st.sidebar.checkbox("Include index", value=False)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    last_step = st.session_state.history[-1]["name"] if st.session_state.history else "export"

    csv_data = df.to_csv(index=include_index).encode("utf-8")
    st.sidebar.download_button(
        label="Download CSV",
        data=csv_data,
        file_name=f"reshape_{last_step}_{timestamp}.csv",
        mime="text/csv",
    )

    buffer = BytesIO()
    df.to_excel(buffer, index=include_index, engine="openpyxl")
    st.sidebar.download_button(
        label="Download XLSX",
        data=buffer.getvalue(),
        file_name=f"reshape_{last_step}_{timestamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def code_snippet_for_last_step() -> str:
    if not st.session_state.history:
        return "# No steps yet"
    step = st.session_state.history[-1]
    name = step["name"]
    params = step["params"]
    code_map = {
        "Loaded": f"df = pd.read_{params.get('source', 'csv')}(...)",
        "Fill down": (
            "for col in columns:\n"
            "    df[col] = df[col].replace('', pd.NA)\n"
            "    df[col] = df[col].ffill()"
        ),
        "Make longer": (
            f"df = pd.melt(df, id_vars={params.get('id_cols')}, "
            f"value_vars={params.get('value_cols')}, "
            f"var_name='{params.get('name_col')}', value_name='{params.get('value_col')}')"
        ),
        "Make wider": (
            f"df = df.pivot_table(index={params.get('index_cols')}, "
            f"columns='{params.get('column_source')}', "
            f"values='{params.get('value_source')}', "
            f"aggfunc={repr(params.get('aggfunc'))})"
        ),
        "Transpose": "df = df.transpose()",
        "Unmerge fill": "df = fill_merged_ranges(workbook, sheet_name, header_row)",
        "Rename columns": "df = df.rename(columns={...})",
    }
    return code_map.get(name, "# Code preview not available for this step")


def main() -> None:
    st.set_page_config(page_title="Reshape Wizard", layout="wide")
    init_state()

    st.title("Reshape Wizard (MVP)")
    st.caption("Offline-friendly reshaping tool with deterministic transformations and preview-first workflow.")

    # --- Sidebar: Import ---
    st.sidebar.header("Import")
    uploaded = st.sidebar.file_uploader("Upload CSV or XLSX", type=["csv", "xlsx", "xlsm", "xls"])
    selected_sheet = None
    header_row = st.sidebar.number_input("Header row (1 = first)", min_value=1, max_value=20, value=1, step=1)
    unmerge_opt = st.sidebar.checkbox("Unmerge cells and fill values (preview/apply in Preprocess tab)", value=False)
    ask_before_fix = st.sidebar.checkbox("Ask before fixing column names", value=False)

    file_bytes: Optional[bytes] = None
    if uploaded is not None:
        file_bytes = uploaded.getvalue()
        ext = uploaded.name.lower().split(".")[-1]
        if ext in ("xlsx", "xlsm", "xls"):
            try:
                sheets = get_sheet_names(file_bytes)
                selected_sheet = st.sidebar.selectbox("Sheet", sheets)
            except Exception as exc:
                st.sidebar.error(f"Could not read sheets: {exc}")

        if st.sidebar.button("Load file", use_container_width=True):
            try:
                df, merge_info = load_dataframe(
                    file_bytes=file_bytes,
                    filename=uploaded.name,
                    sheet_name=selected_sheet,
                    header_row=header_row,
                    unmerge_fill=False,  # unmerge handled in Preprocess tab
                )
                st.session_state.upload_cache = {
                    "bytes": file_bytes,
                    "name": uploaded.name,
                    "ext": ext,
                    "sheet": selected_sheet,
                    "header_row": header_row,
                }
                safe_cols, fixed, mapping = sanitise_columns(list(df.columns))
                if ask_before_fix and fixed:
                    with st.sidebar.form("confirm_columns_load"):
                        st.write("Some column names were empty, Unnamed, or duplicated. Edit before applying.")
                        user_cols = [
                            st.text_input(f"{orig or '(empty)'}", value=suggested, key=f"colfix_load_{i}")
                            for i, (orig, suggested) in enumerate(mapping)
                        ]
                        submitted = st.form_submit_button("Apply column names")
                    if not submitted:
                        st.sidebar.info("Confirm column names to continue.")
                        st.stop()
                    trimmed = [c.strip() for c in user_cols]
                    if any(not c for c in trimmed):
                        st.sidebar.error("Column names cannot be empty.")
                        st.stop()
                    if len(trimmed) != len(set(trimmed)):
                        st.sidebar.error("Column names must be unique.")
                        st.stop()
                    df.columns = trimmed
                else:
                    df.columns = safe_cols
                    if fixed:
                        st.sidebar.info("Some columns had empty or duplicate names and were auto-renamed.")
                st.session_state.original_df = df.copy()
                st.session_state.current_df = df.copy()
                st.session_state.history = []
                add_history_step(
                    "Loaded",
                    {
                        "source": ext,
                        "header_row": header_row,
                        "sheet": selected_sheet,
                        "unmerge_on_load": False,
                        "columns_fixed": fixed,
                    },
                    df,
                    before_shape=df.shape,
                )
                st.session_state.last_merge_fill = merge_info
                clear_preview()
                if df.shape[0] > LARGE_ROW_THRESHOLD:
                    st.sidebar.warning(
                        f"Large dataset ({df.shape[0]} rows). Preview limited to {preview_rows(df)} rows."
                    )
                if unmerge_opt and ext in ("xlsx", "xlsm", "xls"):
                    st.sidebar.info("Unmerge is ready in Preprocess tab. Preview before applying.")
            except Exception as exc:
                st.sidebar.error(f"Load failed: {exc}")

    st.sidebar.markdown("---")
    st.sidebar.subheader("Paste table from Excel")
    pasted_text = st.sidebar.text_area("Paste your table here (Ctrl+V)", height=150)
    if st.sidebar.button("Load pasted table", use_container_width=True):
        try:
            df = load_from_clipboard(pasted_text)
            if df.shape[1] == 1:
                st.sidebar.warning("Only one column detected; check that the table is tab-separated.")
            if df.shape[0] > LARGE_ROW_THRESHOLD:
                st.sidebar.warning(f"Pasted table has {df.shape[0]} rows; previews will be limited.")
            safe_cols, fixed, mapping = sanitise_columns(list(df.columns))
            if ask_before_fix and fixed:
                with st.sidebar.form("confirm_columns_clip"):
                    st.write("Some column names were empty, Unnamed, or duplicated. Edit before applying.")
                    user_cols = [
                        st.text_input(f"{orig or '(empty)'}", value=suggested, key=f"colfix_clip_{i}")
                        for i, (orig, suggested) in enumerate(mapping)
                    ]
                    submitted = st.form_submit_button("Apply column names")
                if not submitted:
                    st.sidebar.info("Confirm column names to continue.")
                    st.stop()
                trimmed = [c.strip() for c in user_cols]
                if any(not c for c in trimmed):
                    st.sidebar.error("Column names cannot be empty.")
                    st.stop()
                if len(trimmed) != len(set(trimmed)):
                    st.sidebar.error("Column names must be unique.")
                    st.stop()
                df.columns = trimmed
            else:
                df.columns = safe_cols
                if fixed:
                    st.sidebar.info("Some columns had empty or duplicate names and were auto-renamed.")
            st.session_state.upload_cache = {}
            st.session_state.original_df = df.copy()
            st.session_state.current_df = df.copy()
            st.session_state.history = []
            add_history_step(
                "Loaded from clipboard",
                {"rows": df.shape[0], "cols": df.shape[1], "columns_fixed": fixed},
                df,
                before_shape=df.shape,
            )
            clear_preview()
            st.sidebar.success("Pasted table loaded.")
        except Exception as exc:
            st.sidebar.error(str(exc))

    # --- Sidebar: Export & History ---
    export_buttons(st.session_state.current_df)
    render_history_sidebar()

    current_df = st.session_state.current_df
    df_hash = df_fingerprint(current_df) if current_df is not None else ""
    auto_allowed = current_df is not None and current_df.shape[0] <= LARGE_ROW_THRESHOLD

    # --- Main area ---
    tabs = st.tabs(["Preview", "Preprocess", "Make longer", "Make wider", "Split combined", "Transpose", "Reproducible code"])

    with tabs[0]:
        show_before_after(current_df, current_df, {"after_shape": current_df.shape if current_df is not None else (0, 0)}, None, "Current data")
        st.markdown("###### Copy table")
        copy_to_clipboard_button(st.session_state.preview_df, current_df, "Copy table to clipboard", "preview_tab")

    # --- Preprocess tab (fill-down + unmerge option) ---
    with tabs[1]:
        st.subheader("Preprocess")
        if current_df is None:
            st.info("Load data first.")
        else:
            fill_cols = st.multiselect("Columns to fill down", current_df.columns.tolist(), key="fill_cols")
            treat_empty = st.checkbox("Treat empty strings as missing", value=True, key="treat_empty")
            auto_update = st.checkbox(
                "Auto-update preview on parameter change",
                value=auto_allowed,
                key="auto_fill",
                disabled=not auto_allowed,
            )
            params = {
                "columns": fill_cols,
                "treat_empty": treat_empty,
            }
            params_hash = f"fill_down|{df_hash}|{json.dumps(params, sort_keys=True)}"
            trigger_preview(current_df, "fill_down", params, auto_update, params_hash, df_hash)
            preview_df, preview_meta, preview_err = get_preview_state(params_hash)
            apply_disabled = (
                preview_err is not None
                or preview_df is None
            )
            if st.button("Apply fill-down", key="apply_fill", disabled=apply_disabled):
                try:
                    result_df, meta = _compute_preview(current_df, "fill_down", params)
                    before_shape = current_df.shape
                    st.session_state.current_df = result_df
                    add_history_step(
                        "Fill down",
                        {**params, "warnings": meta.get("warnings", []), "diagnostics": meta.get("diagnostics", {})},
                        result_df,
                        before_shape,
                    )
                    clear_preview()
                    st.success("Fill-down applied.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"Fill-down failed: {exc}")

            show_before_after(current_df, preview_df, preview_meta, preview_err, "After (preview)")
            copy_to_clipboard_button(preview_df, current_df, "Copy table to clipboard", "preprocess_tab")

            st.markdown("---")
            st.subheader("Unmerge cells and fill (XLSX only)")
            cache = st.session_state.upload_cache
            if cache.get("ext") not in ("xlsx", "xlsm", "xls"):
                st.caption("Load an Excel file to enable unmerge preview.")
            else:
                sheet_choice = st.selectbox("Sheet for unmerge", cache.get("sheet"), key="sheet_unmerge")
                header_choice = st.number_input(
                    "Header row (1 = first)", min_value=1, max_value=20, value=cache.get("header_row", 1), step=1, key="header_unmerge"
                )
                unmerge_params = {
                    "cache_key": cache.get("name"),
                    "sheet_name": sheet_choice,
                    "header_row": int(header_choice),
                }
                unmerge_hash = f"unmerge|{json.dumps(unmerge_params, sort_keys=True)}"
                trigger_preview(current_df, "unmerge_fill", unmerge_params, auto_update, unmerge_hash, df_hash)
                preview_df, preview_meta, preview_err = get_preview_state(unmerge_hash)
                apply_disabled_unmerge = (
                    preview_err is not None
                    or preview_df is None
                )
                if st.button("Apply unmerge+fill", key="apply_unmerge", disabled=apply_disabled_unmerge):
                    try:
                        result_df, meta = _compute_preview(current_df, "unmerge_fill", unmerge_params)
                        before_shape = current_df.shape
                        st.session_state.current_df = result_df
                        add_history_step(
                            "Unmerge fill",
                            {**unmerge_params, "warnings": meta.get("warnings", []), "diagnostics": meta.get("diagnostics", {})},
                            result_df,
                            before_shape,
                        )
                        clear_preview()
                        st.success("Unmerge + fill applied.")
                        st.rerun()
                    except Exception as exc:
                        st.error(f"Unmerge failed: {exc}")
                show_before_after(current_df, preview_df, preview_meta, preview_err, "After (preview)")
                copy_to_clipboard_button(preview_df, current_df, "Copy table to clipboard", "unmerge_tab")

            st.markdown("---")
            st.subheader("Rename columns")
            with st.form("rename_columns_form"):
                new_names = [
                    st.text_input(f"{col}", value=col, key=f"rename_{i}")
                    for i, col in enumerate(current_df.columns)
                ]
                submitted = st.form_submit_button("Apply renames")
            if submitted:
                trimmed = [c.strip() for c in new_names]
                if any(not c for c in trimmed):
                    st.error("Column names cannot be empty.")
                elif len(trimmed) != len(set(trimmed)):
                    st.error("Column names must be unique.")
                else:
                    before_shape = current_df.shape
                    mapping = dict(zip(current_df.columns, trimmed))
                    renamed_df = current_df.rename(columns=mapping)
                    safe_cols, fixed, _ = sanitise_columns(list(renamed_df.columns))
                    if len(safe_cols) != len(set(safe_cols)):
                        st.error("Column names must be unique after sanitisation.")
                    else:
                        renamed_df.columns = safe_cols
                    st.session_state.current_df = renamed_df
                    add_history_step(
                        "Rename columns",
                        {"mapping": mapping},
                        renamed_df,
                        before_shape,
                    )
                    clear_preview()
                    st.success("Columns renamed.")
                    st.rerun()
            copy_to_clipboard_button(st.session_state.preview_df, current_df, "Copy table to clipboard", "rename_tab")

    # --- Make longer ---
    with tabs[2]:
        st.subheader("Make longer (melt)")
        df = st.session_state.current_df
        if df is None:
            st.info("Load data first.")
        else:
            columns = df.columns.tolist()
            id_cols = st.multiselect("ID columns (keep)", columns, key="id_cols_longer")
            value_default = [c for c in columns if c not in id_cols]
            value_cols = st.multiselect("Columns to stack (values)", columns, default=value_default, key="val_cols_longer")
            name_col = st.text_input("Name column", value="variable")
            value_col = st.text_input("Value column", value="value")

            auto_update = st.checkbox(
                "Auto-update preview on parameter change", value=auto_allowed, key="auto_longer", disabled=not auto_allowed
            )
            params = {
                "id_cols": id_cols,
                "value_cols": value_cols,
                "name_col": name_col,
                "value_col": value_col,
            }
            params_hash = f"longer|{df_hash}|{json.dumps(params, sort_keys=True)}"
            trigger_preview(df, "longer", params, auto_update, params_hash, df_hash)
            preview_df, preview_meta, preview_err = get_preview_state(params_hash)

            apply_disabled = (
                preview_err is not None
                or preview_df is None
            )
            if st.button("Apply longer", key="apply_longer", disabled=apply_disabled):
                try:
                    result_df, meta = _compute_preview(df, "longer", params)
                    before_shape = df.shape
                    st.session_state.current_df = result_df
                    add_history_step(
                        "Make longer",
                        {**params, "warnings": meta.get("warnings", []), "diagnostics": meta.get("diagnostics", {})},
                        result_df,
                        before_shape,
                    )
                    clear_preview()
                    st.success("Longer applied.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"Apply failed: {exc}")

            show_before_after(df, preview_df, preview_meta, preview_err, "After (preview)")
            copy_to_clipboard_button(preview_df, df, "Copy table to clipboard", "longer_tab")

    # --- Make wider ---
    with tabs[3]:
        st.subheader("Make wider (pivot)")
        df = st.session_state.current_df
        if df is None:
            st.info("Load data first.")
        else:
            columns = df.columns.tolist()

            index_cols = st.multiselect("Index columns (rows)", columns, key="index_cols_wider")
            combined_parts = st.multiselect("Column name parts (optional, to combine)", columns, key="combined_parts_wider")
            sep = st.text_input("Separator", value="_", key="combined_sep")
            column_source = st.selectbox("Column names from (used if no parts selected)", columns, key="col_source_wider")
            value_source = st.selectbox("Values from", columns, key="val_source_wider")
            agg_options = ["first", "mean", "sum", "count", "min", "max"]

            try:
                work_df_tmp, column_used = build_wider_work_df(df, combined_parts, sep, column_source)
                duplicates = work_df_tmp.duplicated(subset=index_cols + [column_used]).any() if column_used else False
            except Exception as exc:
                duplicates = False
                st.error(f"Preview setup error: {exc}")
            if duplicates:
                st.warning("Duplicates detected in index/column combinations. Please choose an aggregation.")
            aggfunc = st.selectbox("Aggregation (required if duplicates)", [""] + agg_options, index=0 if not duplicates else 1, key="agg_wider")
            aggfunc = aggfunc or None
            reset_idx = st.checkbox("Reset index after pivot", value=True, key="reset_idx_wider")

            auto_update = st.checkbox(
                "Auto-update preview on parameter change", value=auto_allowed, key="auto_wider", disabled=not auto_allowed
            )
            params = {
                "index_cols": index_cols,
                "column_source": column_source,
                "value_source": value_source,
                "aggfunc": aggfunc,
                "reset_index": reset_idx,
                "combined_parts": combined_parts,
                "sep": sep,
            }
            params_hash = f"wider|{df_hash}|{json.dumps(params, sort_keys=True)}"
            trigger_preview(df, "wider", params, auto_update, params_hash, df_hash)
            preview_df, preview_meta, preview_err = get_preview_state(params_hash)

            apply_disabled = (
                preview_err is not None
                or preview_df is None
            )
            if st.button("Apply wider", key="apply_wider", disabled=apply_disabled):
                try:
                    result_df, meta = _compute_preview(df, "wider", params)
                    before_shape = df.shape
                    st.session_state.current_df = result_df
                    add_history_step(
                        "Make wider",
                        {**params, "warnings": meta.get("warnings", []), "diagnostics": meta.get("diagnostics", {})},
                        result_df,
                        before_shape,
                    )
                    clear_preview()
                    st.success("Wider applied.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"Apply failed: {exc}")

            show_before_after(df, preview_df, preview_meta, preview_err, "After (preview)")
            copy_to_clipboard_button(preview_df, df, "Copy table to clipboard", "wider_tab")
            if index_cols and column_source:
                st.caption(f"Key uniqueness check: {'duplicates found' if duplicates else 'all unique'}")

    # --- Split combined columns back to long ---
    with tabs[4]:
        st.subheader("Split combined columns (reverse)")
        df = st.session_state.current_df
        if df is None:
            st.info("Load data first.")
        else:
            columns = df.columns.tolist()
            cols_to_split = st.multiselect("Columns to split", columns, key="split_cols")
            sep_split = st.text_input("Separator in column names", value="_", key="split_sep")
            new_col1 = st.text_input("New column 1 name", value="part1", key="split_new1")
            new_col2 = st.text_input("New column 2 name", value="part2", key="split_new2")
            value_col = st.text_input("Value column name", value="value", key="split_value_col")
            default_ids = [c for c in columns if c not in cols_to_split]
            id_cols = st.multiselect("ID columns to keep", columns, default=default_ids, key="split_id_cols")

            auto_update = st.checkbox(
                "Auto-update preview on parameter change", value=auto_allowed, key="auto_split", disabled=not auto_allowed
            )
            params = {
                "cols_to_split": cols_to_split,
                "sep": sep_split,
                "new_col1": new_col1,
                "new_col2": new_col2,
                "value_col": value_col,
                "id_cols": id_cols,
            }
            params_hash = f"split_combined|{df_hash}|{json.dumps(params, sort_keys=True)}"
            trigger_preview(df, "split_combined", params, auto_update, params_hash, df_hash)
            preview_df, preview_meta, preview_err = get_preview_state(params_hash)

            apply_disabled = (
                preview_err is not None
                or preview_df is None
            )
            if st.button("Apply split", key="apply_split", disabled=apply_disabled):
                try:
                    result_df, meta = _compute_preview(df, "split_combined", params)
                    before_shape = df.shape
                    st.session_state.current_df = result_df
                    add_history_step(
                        "Split combined",
                        {**params, "warnings": meta.get("warnings", []), "diagnostics": meta.get("diagnostics", {})},
                        result_df,
                        before_shape,
                    )
                    clear_preview()
                    st.success("Split applied.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"Apply failed: {exc}")

            show_before_after(df, preview_df, preview_meta, preview_err, "After (preview)")
            copy_to_clipboard_button(preview_df, df, "Copy table to clipboard", "split_tab")

    # --- Transpose ---
    with tabs[5]:
        st.subheader("Transpose")
        df = st.session_state.current_df
        if df is None:
            st.info("Load data first.")
        else:
            first_row_header = st.checkbox("First row becomes headers after transpose", value=True, key="first_row_header")
            first_col_index = st.checkbox("First column becomes row names before transpose", value=False, key="first_col_index")

            auto_update = st.checkbox(
                "Auto-update preview on parameter change", value=auto_allowed, key="auto_transpose", disabled=not auto_allowed
            )
            params = {"first_row_header": first_row_header, "first_col_index": first_col_index}
            params_hash = f"transpose|{df_hash}|{json.dumps(params, sort_keys=True)}"
            trigger_preview(df, "transpose", params, auto_update, params_hash, df_hash)
            preview_df, preview_meta, preview_err = get_preview_state(params_hash)

            apply_disabled = (
                preview_err is not None
                or preview_df is None
            )
            if st.button("Apply transpose", key="apply_transpose", disabled=apply_disabled):
                try:
                    result_df, meta = _compute_preview(df, "transpose", params)
                    before_shape = df.shape
                    st.session_state.current_df = result_df
                    add_history_step(
                        "Transpose",
                        {**params, "warnings": meta.get("warnings", []), "diagnostics": meta.get("diagnostics", {})},
                        result_df,
                        before_shape,
                    )
                    clear_preview()
                    st.success("Transpose applied.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"Apply failed: {exc}")

            show_before_after(df, preview_df, preview_meta, preview_err, "After (preview)")
            copy_to_clipboard_button(preview_df, df, "Copy table to clipboard", "transpose_tab")

    with tabs[6]:
        st.subheader("Reproducible code")
        st.code(code_snippet_for_last_step(), language="python")


if __name__ == "__main__":
    main()
