from __future__ import annotations
import io
import os
from typing import Optional, Tuple, List, Dict, Any, BinaryIO
import pandas as pd
from .clean_columns import sanitise_columns


class FileContext:
    def __init__(self, path: Optional[str] = None, ext: str = "", sheet_name: Optional[str] = None, header_row: int = 0, bytes_data: bytes | None = None):
        self.path = path
        self.ext = ext
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.bytes_data = bytes_data


def read_csv_with_fallback(path: str, encoding_try: Tuple[str, ...] = ("utf-8", "cp1251")) -> pd.DataFrame:
    last_err: Exception | None = None
    for enc in encoding_try:
        try:
            df = pd.read_csv(path, encoding=enc)
            return sanitise_columns(df)[0]
        except Exception as exc:  # pragma: no cover - best effort
            last_err = exc
            continue
    if last_err:
        raise last_err
    raise RuntimeError("Could not read CSV")


def load_xlsx(path: str, sheet_name: Optional[str] = None, header_row: int = 0, unmerge_fill: bool = False) -> pd.DataFrame:
    if unmerge_fill:
        df, _ = unmerge_fill_xlsx(path_or_bytes=path, sheet_name=sheet_name, header_row=header_row)
    else:
        df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)
    return sanitise_columns(df)[0]


def parse_clipboard_tsv(text: str) -> pd.DataFrame:
    buf = io.StringIO(text)
    df = pd.read_csv(buf, sep="\t")
    if df.columns.isnull().any():
        df.columns = [f"col{idx+1}" for idx in range(len(df.columns))]
    return sanitise_columns(df)[0]


def get_sheet_names(path: str) -> List[str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    return wb.sheetnames


def unmerge_fill_xlsx(path_or_bytes: str | bytes | BinaryIO, sheet_name: Optional[str] = None, header_row: int = 0) -> tuple[pd.DataFrame, Dict[str, Any]]:
    from openpyxl import load_workbook
    if isinstance(path_or_bytes, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(path_or_bytes))
    elif hasattr(path_or_bytes, "read"):
        wb = load_workbook(path_or_bytes)
    else:
        wb = load_workbook(path_or_bytes)
    ws = wb[sheet_name] if sheet_name else wb.active
    filled_ranges = 0
    for merged in list(ws.merged_cells.ranges):
        filled_ranges += 1
        top_left = merged.min_row, merged.min_col
        value = ws.cell(*top_left).value
        for row in ws.iter_rows(min_row=merged.min_row, max_row=merged.max_row, min_col=merged.min_col, max_col=merged.max_col):
            for cell in row:
                cell.value = value
        ws.unmerge_cells(str(merged))
    data = [[cell.value for cell in row] for row in ws.iter_rows(values_only=True)]
    df = pd.DataFrame(data)
    df.columns = df.iloc[header_row]
    df = df.drop(index=list(range(header_row + 1)))
    df, mapping, fixed = sanitise_columns(df)
    meta = {"filled_ranges": filled_ranges, "mapping": mapping, "fixed": fixed}
    return df, meta


def load_file_via_context(path: str, sheet_name: Optional[str] = None, header_row: int = 0, unmerge_fill: bool = False) -> tuple[pd.DataFrame, FileContext]:
    ext = os.path.splitext(path)[1].lower().lstrip('.')
    if ext == "csv":
        df = read_csv_with_fallback(path)
        ctx = FileContext(path=path, ext=ext, sheet_name=None, header_row=header_row)
        return df, ctx
    if ext in ("xlsx", "xlsm", "xlsb"):
        if unmerge_fill:
            df, meta = unmerge_fill_xlsx(path, sheet_name, header_row)
        else:
            df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)
            df, _, _ = sanitise_columns(df)
        with open(path, "rb") as f:
            bytes_data = f.read()
        ctx = FileContext(path=path, ext=ext, sheet_name=sheet_name, header_row=header_row, bytes_data=bytes_data)
        return df, ctx
    raise ValueError("Unsupported file type")
